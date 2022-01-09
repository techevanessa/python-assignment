
def change(to_change, change_to, file):
    import csv
    with open(file, 'r') as f:
        reader = csv.reader(f)
        data = list(reader)
        for row in data:
            if(str(row[1]).endswith(to_change)):
                row[1] = row[1].replace(to_change, change_to)

    with open('clean.csv', 'w') as f:
        import csv
        writer = csv.writer(f)
        writer.writerows(data)


def changexlsx():
    from openpyxl import load_workbook

    wb = load_workbook('employeedada.xlsx')
    wb = wb.active
    sheet = wb['sheet1']

    for i in range (2, sheet.max_row1):
        cell = sheet.cell(1.2)
        if 'helpinghands.cm'in cell.value:
            update = (cell.value).replace('helpinghands.cm','handsinhands.org')
            sheet.cell(1,2).value = update
    wb.save('updated_emails.xlsx')

to_change ="@helpinghands.cm"
change_to = "@handinhands.org"
file = 'employeedata.csv'


#change(to_change, change_to, file)
changexlsx()